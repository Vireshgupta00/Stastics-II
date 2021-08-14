# Title:Reading data from .xlsx file (Excel File).
# 
# Description: A python program: This code extracts the datate from an excel file 
#              and print it on the CMD or IDLE shell.
# 
# Copyright: Viresh Gupta Â© 2021 
# 
# Author: Viresh Gupta
# 
# Version: 1.00	     2021 Baseline

import xlrd 
import openpyxl

xlrd.xlsx.ensure_elementtree_imported(False, None)
xlrd.xlsx.Element_has_iter = True
 
#The file to be opened 
loc = (r"marks.xlsx")
  
# To open Workbook 
wb = xlrd.open_workbook(loc)
worksheet = wb.sheet_by_name('Sheet1') 
worksheet = wb.sheet_by_index(0) 

# To read and print the file contents
i=0
while (i<11):
 sn = worksheet.cell(i, 0)
 rn = worksheet.cell(i, 1)
 mrks = worksheet.cell(i, 2)
 print(" ", str(sn.value), " , ", str(rn.value), " , ",str(mrks.value)) 
 i=i+1